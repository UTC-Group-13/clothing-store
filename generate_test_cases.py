"""
Script tạo file Excel kịch bản kiểm thử chức năng Đăng Nhập
API: POST /api/auth/login
Project: Clothing Store E-commerce API
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime

# ─────────────────────────────────────────────────────────────────────────────
# DỮ LIỆU KỊCH BẢN KIỂM THỬ
# ─────────────────────────────────────────────────────────────────────────────
TEST_CASES = [
    # ── NHÓM 1: KIỂM THỬ ĐẦU VÀO HỢP LỆ ────────────────────────────────────
    {
        "id": "TC_LOGIN_001",
        "group": "Đăng nhập thành công",
        "title": "Đăng nhập với username & password hợp lệ (user thường)",
        "precondition": "Tài khoản tồn tại trong DB: username='john_doe', password='password123'",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "200 OK",
        "expected_response": "success=true, accessToken không rỗng, tokenType='Bearer', role='USER'",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "Happy path — luồng chính",
    },
    {
        "id": "TC_LOGIN_002",
        "group": "Đăng nhập thành công",
        "title": "Đăng nhập với tài khoản ADMIN hợp lệ",
        "precondition": "Tài khoản ADMIN tồn tại: username='admin', password='admin123'",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "admin", "password": "admin123"}',
        "expected_status": "200 OK",
        "expected_response": "success=true, accessToken không rỗng, role='ADMIN'",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "Xác nhận role ADMIN trong token",
    },
    {
        "id": "TC_LOGIN_003",
        "group": "Đăng nhập thành công",
        "title": "Đăng nhập — password có ký tự đặc biệt hợp lệ",
        "precondition": "Tài khoản tồn tại: username='user_special', password='P@ss#w0rd!2024'",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "user_special", "password": "P@ss#w0rd!2024"}',
        "expected_status": "200 OK",
        "expected_response": "success=true, accessToken không rỗng",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "Kiểm tra password có ký tự đặc biệt: @, #, !",
    },
    {
        "id": "TC_LOGIN_004",
        "group": "Đăng nhập thành công",
        "title": "Đăng nhập — JWT token phải có thời hạn 24 giờ",
        "precondition": "Tài khoản tồn tại hợp lệ",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "200 OK",
        "expected_response": "accessToken decode được, exp - iat = 86400 giây (24h)",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "Giải mã JWT và kiểm tra trường exp",
    },
    {
        "id": "TC_LOGIN_005",
        "group": "Đăng nhập thành công",
        "title": "Đăng nhập nhiều lần liên tiếp — mỗi lần cấp token mới",
        "precondition": "Tài khoản hợp lệ",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "200 OK",
        "expected_response": "Mỗi lần trả về accessToken khác nhau (timestamp iat khác nhau)",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "Gọi 2 lần, so sánh 2 token",
    },

    # ── NHÓM 2: VALIDATION ĐẦU VÀO ──────────────────────────────────────────
    {
        "id": "TC_LOGIN_006",
        "group": "Validation đầu vào",
        "title": "Username để trống",
        "precondition": "Không cần tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "", "password": "password123"}',
        "expected_status": "400 Bad Request",
        "expected_response": "success=false, message chứa 'Username không được để trống'",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "@NotBlank validation trên LoginRequest.username",
    },
    {
        "id": "TC_LOGIN_007",
        "group": "Validation đầu vào",
        "title": "Password để trống",
        "precondition": "Không cần tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": ""}',
        "expected_status": "400 Bad Request",
        "expected_response": "success=false, message chứa 'Password không được để trống'",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "@NotBlank validation trên LoginRequest.password",
    },
    {
        "id": "TC_LOGIN_008",
        "group": "Validation đầu vào",
        "title": "Cả username và password đều để trống",
        "precondition": "Không cần tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "", "password": ""}',
        "expected_status": "400 Bad Request",
        "expected_response": "success=false, trả về 2 lỗi validation",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "GlobalExceptionHandler phải trả về list lỗi đầy đủ",
    },
    {
        "id": "TC_LOGIN_009",
        "group": "Validation đầu vào",
        "title": "Thiếu trường username trong request body",
        "precondition": "Không cần tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"password": "password123"}',
        "expected_status": "400 Bad Request",
        "expected_response": "success=false, validation error username",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "Thiếu field hoàn toàn — khác với empty string",
    },
    {
        "id": "TC_LOGIN_010",
        "group": "Validation đầu vào",
        "title": "Thiếu trường password trong request body",
        "precondition": "Không cần tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe"}',
        "expected_status": "400 Bad Request",
        "expected_response": "success=false, validation error password",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "Thiếu field hoàn toàn",
    },
    {
        "id": "TC_LOGIN_011",
        "group": "Validation đầu vào",
        "title": "Request body hoàn toàn rỗng {}",
        "precondition": "Không cần tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": "{}",
        "expected_status": "400 Bad Request",
        "expected_response": "success=false, 2 lỗi validation",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "",
    },
    {
        "id": "TC_LOGIN_012",
        "group": "Validation đầu vào",
        "title": "Request body không phải JSON (plain text)",
        "precondition": "Không cần tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": "username=john_doe&password=123456",
        "expected_status": "400 Bad Request",
        "expected_response": "HttpMessageNotReadableException — không parse được body",
        "actual_response": "",
        "result": "",
        "priority": "Thấp",
        "note": "Content-Type sai (form-urlencoded thay vì application/json)",
    },
    {
        "id": "TC_LOGIN_013",
        "group": "Validation đầu vào",
        "title": "Username chỉ có khoảng trắng (whitespace)",
        "precondition": "Không cần tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "   ", "password": "password123"}',
        "expected_status": "400 Bad Request",
        "expected_response": "success=false — @NotBlank bắt whitespace-only",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "@NotBlank trim whitespace trước khi check",
    },
    {
        "id": "TC_LOGIN_014",
        "group": "Validation đầu vào",
        "title": "Username null (giá trị null trong JSON)",
        "precondition": "Không cần tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": null, "password": "password123"}',
        "expected_status": "400 Bad Request",
        "expected_response": "success=false — @NotBlank bắt null",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "",
    },

    # ── NHÓM 3: SAI THÔNG TIN ĐĂNG NHẬP ─────────────────────────────────────
    {
        "id": "TC_LOGIN_015",
        "group": "Sai thông tin đăng nhập",
        "title": "Username đúng, password sai",
        "precondition": "Tài khoản tồn tại: username='john_doe'",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "wrongpassword"}',
        "expected_status": "401 Unauthorized",
        "expected_response": "success=false, message='Sai tài khoản hoặc mật khẩu' (hoặc tương đương)",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "BadCredentialsException từ AuthenticationManager",
    },
    {
        "id": "TC_LOGIN_016",
        "group": "Sai thông tin đăng nhập",
        "title": "Username không tồn tại trong hệ thống",
        "precondition": "Username 'unknown_user_xyz' không có trong DB",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "unknown_user_xyz", "password": "password123"}',
        "expected_status": "401 Unauthorized",
        "expected_response": "success=false, message báo sai thông tin",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "Không nên tiết lộ 'username không tồn tại' (bảo mật)",
    },
    {
        "id": "TC_LOGIN_017",
        "group": "Sai thông tin đăng nhập",
        "title": "Cả username và password đều sai",
        "precondition": "Không tài khoản nào khớp",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "fake_user", "password": "fake_pass"}',
        "expected_status": "401 Unauthorized",
        "expected_response": "success=false",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "",
    },
    {
        "id": "TC_LOGIN_018",
        "group": "Sai thông tin đăng nhập",
        "title": "Username đúng nhưng thừa khoảng trắng đầu/cuối",
        "precondition": "Tài khoản 'john_doe' tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": " john_doe ", "password": "password123"}',
        "expected_status": "401 Unauthorized",
        "expected_response": "success=false — username ' john_doe ' khác 'john_doe'",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "Hệ thống không tự trim username khi auth",
    },
    {
        "id": "TC_LOGIN_019",
        "group": "Sai thông tin đăng nhập",
        "title": "Username đúng nhưng sai hoa/thường (case sensitive)",
        "precondition": "Tài khoản 'john_doe' tồn tại (chữ thường)",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "JOHN_DOE", "password": "password123"}',
        "expected_status": "401 Unauthorized",
        "expected_response": "success=false — username không khớp (case sensitive)",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "MySQL/Spring Security username case sensitive",
    },
    {
        "id": "TC_LOGIN_020",
        "group": "Sai thông tin đăng nhập",
        "title": "Password sai 1 ký tự",
        "precondition": "Tài khoản hợp lệ, password='password123'",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password124"}',
        "expected_status": "401 Unauthorized",
        "expected_response": "success=false",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "BCrypt so sánh chính xác từng ký tự",
    },

    # ── NHÓM 4: BẢO MẬT ─────────────────────────────────────────────────────
    {
        "id": "TC_LOGIN_021",
        "group": "Bảo mật",
        "title": "SQL Injection trong trường username",
        "precondition": "DB có dữ liệu",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "\' OR \'1\'=\'1", "password": "password"}',
        "expected_status": "401 Unauthorized",
        "expected_response": "success=false — Spring Security/JPA chặn SQL injection",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "Spring Data JPA dùng PreparedStatement — không bị SQL injection",
    },
    {
        "id": "TC_LOGIN_022",
        "group": "Bảo mật",
        "title": "SQL Injection trong trường password",
        "precondition": "DB có dữ liệu",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "\' OR 1=1 --"}',
        "expected_status": "401 Unauthorized",
        "expected_response": "success=false — không bypass được BCrypt",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "BCrypt hash so sánh — không thể bypass bằng SQL injection",
    },
    {
        "id": "TC_LOGIN_023",
        "group": "Bảo mật",
        "title": "XSS trong trường username",
        "precondition": "Không cần tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "<script>alert(1)</script>", "password": "pass"}',
        "expected_status": "401 Unauthorized",
        "expected_response": "success=false — không thực thi script, trả về lỗi auth",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "JSON API không render HTML — XSS không nguy hiểm ở API layer",
    },
    {
        "id": "TC_LOGIN_024",
        "group": "Bảo mật",
        "title": "Password không được trả về trong response",
        "precondition": "Tài khoản hợp lệ",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "200 OK",
        "expected_response": "Response KHÔNG chứa trường 'password' hoặc hash password",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "AuthResponse chỉ chứa: accessToken, tokenType, userId, username, role",
    },
    {
        "id": "TC_LOGIN_025",
        "group": "Bảo mật",
        "title": "Brute force — gọi 10 lần liên tiếp với password sai",
        "precondition": "Tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "wrongpassword"}',
        "expected_status": "401 Unauthorized (mỗi lần)",
        "expected_response": "Hệ thống có thể rate-limit sau N lần sai (nếu đã cấu hình)",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "Hiện tại project chưa có rate limiting — ghi nhận để cải tiến",
    },
    {
        "id": "TC_LOGIN_026",
        "group": "Bảo mật",
        "title": "Password cực dài (>10.000 ký tự) — kiểm tra DoS",
        "precondition": "Không cần tài khoản tồn tại",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "A..." (10.000 ký tự)}',
        "expected_status": "400 Bad Request hoặc 401",
        "expected_response": "Hệ thống không bị timeout/crash khi BCrypt hash chuỗi dài",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "BCrypt có thể tốn nhiều CPU với password dài — cần giới hạn input",
    },

    # ── NHÓM 5: KIỂM THỬ RESPONSE FORMAT ────────────────────────────────────
    {
        "id": "TC_LOGIN_027",
        "group": "Response format",
        "title": "Cấu trúc response success đúng chuẩn ApiResponse<T>",
        "precondition": "Tài khoản hợp lệ",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "200 OK",
        "expected_response": '{"success":true,"message":"Đăng nhập thành công","data":{...},"timestamp":"..."}',
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "Kiểm tra đủ 4 field: success, message, data, timestamp",
    },
    {
        "id": "TC_LOGIN_028",
        "group": "Response format",
        "title": "Trường data.accessToken phải là JWT hợp lệ (3 phần dot-separated)",
        "precondition": "Tài khoản hợp lệ",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "200 OK",
        "expected_response": "accessToken có dạng: xxxxx.yyyyy.zzzzz (base64.base64.base64)",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "Dùng jwt.io để decode và verify",
    },
    {
        "id": "TC_LOGIN_029",
        "group": "Response format",
        "title": "Trường data.tokenType = 'Bearer'",
        "precondition": "Tài khoản hợp lệ",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "200 OK",
        "expected_response": 'data.tokenType = "Bearer"',
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "",
    },
    {
        "id": "TC_LOGIN_030",
        "group": "Response format",
        "title": "Trường data.userId là số nguyên dương",
        "precondition": "Tài khoản hợp lệ",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "200 OK",
        "expected_response": "data.userId là Integer > 0, khớp với ID trong DB",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "",
    },
    {
        "id": "TC_LOGIN_031",
        "group": "Response format",
        "title": "Content-Type response là application/json",
        "precondition": "Tài khoản hợp lệ",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "200 OK",
        "expected_response": "Header: Content-Type: application/json",
        "actual_response": "",
        "result": "",
        "priority": "Trung bình",
        "note": "",
    },

    # ── NHÓM 6: SỬ DỤNG TOKEN SAU KHI ĐĂNG NHẬP ─────────────────────────────
    {
        "id": "TC_LOGIN_032",
        "group": "Sử dụng JWT token",
        "title": "Dùng token hợp lệ gọi API cần xác thực (GET /api/cart)",
        "precondition": "Đã đăng nhập, có accessToken hợp lệ",
        "request_method": "GET",
        "request_url": "/api/cart",
        "request_body": "(không có body — dùng header Authorization: Bearer <token>)",
        "expected_status": "200 OK",
        "expected_response": "Trả về giỏ hàng của user",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "Xác nhận token hoạt động đúng sau login",
    },
    {
        "id": "TC_LOGIN_033",
        "group": "Sử dụng JWT token",
        "title": "Gọi API cần xác thực mà không có token",
        "precondition": "Không có Authorization header",
        "request_method": "GET",
        "request_url": "/api/cart",
        "request_body": "(không có header Authorization)",
        "expected_status": "403 Forbidden hoặc 401 Unauthorized",
        "expected_response": "success=false — JwtAuthenticationFilter chặn",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "",
    },
    {
        "id": "TC_LOGIN_034",
        "group": "Sử dụng JWT token",
        "title": "Dùng token giả mạo (invalid signature)",
        "precondition": "Có token fake",
        "request_method": "GET",
        "request_url": "/api/cart",
        "request_body": "(Header: Authorization: Bearer fake.jwt.token)",
        "expected_status": "403 Forbidden",
        "expected_response": "success=false — JwtService.isTokenValid() = false",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "",
    },
    {
        "id": "TC_LOGIN_035",
        "group": "Sử dụng JWT token",
        "title": "Dùng token hết hạn (expired token)",
        "precondition": "Có token đã hết hạn (exp < now)",
        "request_method": "GET",
        "request_url": "/api/cart",
        "request_body": "(Header: Authorization: Bearer <expired_token>)",
        "expected_status": "403 Forbidden",
        "expected_response": "success=false — ExpiredJwtException",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "Cần đợi 24h hoặc set JWT_EXPIRATION ngắn để test",
    },
    {
        "id": "TC_LOGIN_036",
        "group": "Sử dụng JWT token",
        "title": "Token hợp lệ nhưng truy cập endpoint ADMIN (user thường)",
        "precondition": "Token của user thường (role=USER)",
        "request_method": "GET",
        "request_url": "/api/orders/admin/all",
        "request_body": "(Header: Authorization: Bearer <user_token>)",
        "expected_status": "403 Forbidden",
        "expected_response": "success=false — @PreAuthorize('hasRole(ADMIN)') từ chối",
        "actual_response": "",
        "result": "",
        "priority": "Cao",
        "note": "Kiểm tra phân quyền RBAC",
    },

    # ── NHÓM 7: HTTP METHOD & HEADER ─────────────────────────────────────────
    {
        "id": "TC_LOGIN_037",
        "group": "HTTP Method & Header",
        "title": "Gọi với method GET thay vì POST",
        "precondition": "Không cần tài khoản",
        "request_method": "GET",
        "request_url": "/api/auth/login",
        "request_body": "(không có body)",
        "expected_status": "405 Method Not Allowed",
        "expected_response": "success=false — Spring MVC từ chối GET",
        "actual_response": "",
        "result": "",
        "priority": "Thấp",
        "note": "",
    },
    {
        "id": "TC_LOGIN_038",
        "group": "HTTP Method & Header",
        "title": "Gọi với method PUT thay vì POST",
        "precondition": "Không cần tài khoản",
        "request_method": "PUT",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "405 Method Not Allowed",
        "expected_response": "success=false",
        "actual_response": "",
        "result": "",
        "priority": "Thấp",
        "note": "",
    },
    {
        "id": "TC_LOGIN_039",
        "group": "HTTP Method & Header",
        "title": "Content-Type là text/plain thay vì application/json",
        "precondition": "Không cần tài khoản",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "415 Unsupported Media Type hoặc 400",
        "expected_response": "success=false — Spring không parse JSON",
        "actual_response": "",
        "result": "",
        "priority": "Thấp",
        "note": "Header: Content-Type: text/plain",
    },
    {
        "id": "TC_LOGIN_040",
        "group": "HTTP Method & Header",
        "title": "CORS — gọi từ origin không được phép",
        "precondition": "Không cần tài khoản",
        "request_method": "POST",
        "request_url": "/api/auth/login",
        "request_body": '{"username": "john_doe", "password": "password123"}',
        "expected_status": "403 Forbidden (CORS block ở browser)",
        "expected_response": "Browser block — không có Access-Control-Allow-Origin header",
        "actual_response": "",
        "result": "",
        "priority": "Thấp",
        "note": "Chỉ test được trên browser/curl — postman bỏ qua CORS",
    },
]

# ─────────────────────────────────────────────────────────────────────────────
# MÀU SẮC & STYLE
# ─────────────────────────────────────────────────────────────────────────────
COLORS = {
    "header_bg":    "1F3864",   # Xanh navy đậm
    "header_font":  "FFFFFF",   # Trắng
    "group_1":      "D6E4F0",   # Xanh nhạt — Đăng nhập thành công
    "group_2":      "FFF2CC",   # Vàng nhạt — Validation
    "group_3":      "FCE4D6",   # Cam nhạt — Sai thông tin
    "group_4":      "FFD7D7",   # Đỏ nhạt — Bảo mật
    "group_5":      "E2EFDA",   # Xanh lá nhạt — Response format
    "group_6":      "EAD1DC",   # Hồng nhạt — Sử dụng JWT
    "group_7":      "F4CCFF",   # Tím nhạt — HTTP Method
    "pass":         "C6EFCE",   # Xanh PASS
    "fail":         "FFC7CE",   # Đỏ FAIL
    "skip":         "FFEB9C",   # Vàng SKIP
    "row_alt":      "F8F9FA",   # Xám nhạt (dòng chẵn)
    "priority_cao": "FF0000",   # Đỏ
    "priority_tb":  "FF8C00",   # Cam
    "priority_low": "008000",   # Xanh lá
    "title_bg":     "2F5496",   # Xanh title
}

GROUP_COLORS = {
    "Đăng nhập thành công":  COLORS["group_1"],
    "Validation đầu vào":    COLORS["group_2"],
    "Sai thông tin đăng nhập": COLORS["group_3"],
    "Bảo mật":               COLORS["group_4"],
    "Response format":       COLORS["group_5"],
    "Sử dụng JWT token":     COLORS["group_6"],
    "HTTP Method & Header":  COLORS["group_7"],
}

PRIORITY_COLORS = {
    "Cao":       "FF0000",
    "Trung bình": "FF8C00",
    "Thấp":      "008000",
}


def make_border(style="thin"):
    s = Side(style=style, color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def set_cell(ws, row, col, value, bold=False, font_color="000000",
             bg_color=None, align="left", wrap=True, font_size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Calibri", bold=bold, color=font_color, size=font_size)
    cell.alignment = Alignment(
        horizontal=align, vertical="center",
        wrap_text=wrap
    )
    if bg_color:
        cell.fill = make_fill(bg_color)
    cell.border = make_border()
    return cell


# ─────────────────────────────────────────────────────────────────────────────
# TẠO SHEET 1: TRANG BÌA
# ─────────────────────────────────────────────────────────────────────────────
def create_cover_sheet(wb):
    ws = wb.create_sheet("📋 Trang Bìa", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 50

    # Tiêu đề chính
    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "🧪 KỊCH BẢN KIỂM THỬ CHỨC NĂNG ĐĂNG NHẬP"
    c.font = Font(name="Calibri", bold=True, size=18, color="FFFFFF")
    c.fill = make_fill(COLORS["title_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 50

    # Subtitle
    ws.merge_cells("B3:C3")
    c = ws["B3"]
    c.value = "Clothing Store E-commerce API — POST /api/auth/login"
    c.font = Font(name="Calibri", bold=False, size=12, color="2F5496")
    c.fill = make_fill("DEEAF1")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 25

    info = [
        ("📌 Dự án",       "Clothing Store E-commerce API"),
        ("🔗 Endpoint",    "POST /api/auth/login"),
        ("📅 Ngày tạo",    datetime.now().strftime("%d/%m/%Y")),
        ("👤 Người tạo",   "QA Team"),
        ("🔢 Tổng số TC",  str(len(TEST_CASES))),
        ("🛠️ Môi trường", "http://localhost:8080"),
        ("📦 Version",     "Spring Boot 3.5.11 / Java 21"),
        ("📝 Trạng thái",  "DRAFT"),
    ]

    for i, (label, value) in enumerate(info):
        r = 5 + i
        ws.row_dimensions[r].height = 22
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = Font(name="Calibri", bold=True, size=11, color="1F3864")
        lc.fill = make_fill("DEEAF1")
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = make_border()

        vc = ws.cell(row=r, column=3, value=value)
        vc.font = Font(name="Calibri", size=11, color="000000")
        vc.fill = make_fill("FFFFFF" if i % 2 == 0 else "F2F2F2")
        vc.alignment = Alignment(horizontal="left", vertical="center")
        vc.border = make_border()

    # Ghi chú nhóm test
    r = 14
    ws.merge_cells(f"B{r}:C{r}")
    c = ws[f"B{r}"]
    c.value = "📂 PHÂN NHÓM KIỂM THỬ"
    c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    c.fill = make_fill(COLORS["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 25

    groups = [
        ("TC_LOGIN_001 – 005", "Đăng nhập thành công",      "5 TC", COLORS["group_1"]),
        ("TC_LOGIN_006 – 014", "Validation đầu vào",         "9 TC", COLORS["group_2"]),
        ("TC_LOGIN_015 – 020", "Sai thông tin đăng nhập",    "6 TC", COLORS["group_3"]),
        ("TC_LOGIN_021 – 026", "Bảo mật",                    "6 TC", COLORS["group_4"]),
        ("TC_LOGIN_027 – 031", "Response format",            "5 TC", COLORS["group_5"]),
        ("TC_LOGIN_032 – 036", "Sử dụng JWT token",          "5 TC", COLORS["group_6"]),
        ("TC_LOGIN_037 – 040", "HTTP Method & Header",       "4 TC", COLORS["group_7"]),
    ]
    for i, (tc_range, g_name, count, color) in enumerate(groups):
        row = 15 + i
        ws.row_dimensions[row].height = 22
        for j, val in enumerate([tc_range, g_name, count]):
            c = ws.cell(row=row, column=2 + j if j < 2 else 3, value=val)
            c.fill = make_fill(color)
            c.font = Font(name="Calibri", size=10, bold=(j == 1))
            c.alignment = Alignment(horizontal="left" if j < 2 else "center", vertical="center")
            c.border = make_border()
        # merge cột 2
        if True:
            pass  # individual cells for now

    ws.freeze_panes = None


# ─────────────────────────────────────────────────────────────────────────────
# TẠO SHEET 2: DANH SÁCH TEST CASES
# ─────────────────────────────────────────────────────────────────────────────
def create_testcase_sheet(wb):
    ws = wb.create_sheet("🧪 Test Cases", 1)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"

    # Chiều rộng cột
    col_widths = {
        "A": 14,   # Test ID
        "B": 35,   # Nhóm
        "C": 50,   # Tiêu đề / Mô tả
        "D": 45,   # Điều kiện tiên quyết
        "E": 10,   # Method
        "F": 28,   # URL
        "G": 55,   # Request Body
        "H": 12,   # Expected HTTP Status
        "I": 60,   # Expected Response
        "J": 45,   # Actual Response
        "K": 14,   # Kết quả (PASS/FAIL/SKIP)
        "L": 10,   # Độ ưu tiên
        "M": 35,   # Ghi chú
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── HEADER ──
    headers = [
        "TEST ID", "NHÓM", "MÔ TẢ / TIÊU ĐỀ",
        "ĐIỀU KIỆN TIÊN QUYẾT", "METHOD", "URL",
        "REQUEST BODY", "EXPECTED\nHTTP STATUS",
        "EXPECTED RESPONSE", "ACTUAL RESPONSE",
        "KẾT QUẢ", "ƯU TIÊN", "GHI CHÚ"
    ]

    # Merge title row
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = "🧪  KỊCH BẢN KIỂM THỬ — CHỨC NĂNG ĐĂNG NHẬP  |  POST /api/auth/login  |  Clothing Store API"
    title_cell.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    title_cell.fill = make_fill(COLORS["title_bg"])
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = make_border("medium")
    ws.row_dimensions[1].height = 32

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(name="Calibri", bold=True, size=10, color=COLORS["header_font"])
        cell.fill = make_fill(COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = make_border("medium")
    ws.row_dimensions[2].height = 38

    # ── DỮ LIỆU ──
    for row_idx, tc in enumerate(TEST_CASES, start=3):
        ws.row_dimensions[row_idx].height = 75

        group_color = GROUP_COLORS.get(tc["group"], "FFFFFF")
        base_color = group_color if row_idx % 2 != 0 else COLORS["row_alt"]

        values = [
            tc["id"],
            tc["group"],
            tc["title"],
            tc["precondition"],
            tc["request_method"],
            tc["request_url"],
            tc["request_body"],
            tc["expected_status"],
            tc["expected_response"],
            tc["actual_response"],
            tc["result"],
            tc["priority"],
            tc["note"],
        ]

        for col_idx, value in enumerate(values, start=1):
            col_letter = get_column_letter(col_idx)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(
                horizontal="center" if col_idx in [1, 5, 8, 11, 12] else "left",
                vertical="center",
                wrap_text=True
            )
            cell.border = make_border()
            cell.font = Font(name="Calibri", size=9)

            # Màu nền theo nhóm
            if col_idx == 1:  # Test ID
                cell.fill = make_fill(group_color)
                cell.font = Font(name="Calibri", bold=True, size=9, color="1F3864")
            elif col_idx == 2:  # Nhóm
                cell.fill = make_fill(group_color)
                cell.font = Font(name="Calibri", bold=True, size=9)
            elif col_idx == 5:  # Method
                method_colors = {
                    "POST": "FF8C00", "GET": "1F8C00",
                    "PUT": "2F5496", "DELETE": "C00000",
                    "PATCH": "7030A0"
                }
                m_color = method_colors.get(value, "666666")
                cell.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
                cell.fill = make_fill(m_color)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 8:  # Expected Status
                status_bg = "C6EFCE" if "200" in str(value) else (
                    "FFC7CE" if any(x in str(value) for x in ["4", "5"]) else "FFFFFF"
                )
                cell.fill = make_fill(status_bg)
                cell.font = Font(name="Calibri", bold=True, size=9)
            elif col_idx == 11:  # Kết quả
                result_bg = {
                    "PASS": COLORS["pass"],
                    "FAIL": COLORS["fail"],
                    "SKIP": COLORS["skip"],
                }.get(value, "FFFFFF")
                cell.fill = make_fill(result_bg)
                cell.font = Font(name="Calibri", bold=True, size=10)
            elif col_idx == 12:  # Ưu tiên
                p_color = PRIORITY_COLORS.get(value, "000000")
                cell.font = Font(name="Calibri", bold=True, size=9, color=p_color)
                cell.fill = make_fill("FFFFFF")
            else:
                cell.fill = make_fill(group_color if col_idx in [3] else "FFFFFF")

    # Auto filter
    ws.auto_filter.ref = f"A2:M{len(TEST_CASES) + 2}"


# ─────────────────────────────────────────────────────────────────────────────
# TẠO SHEET 3: THỐNG KÊ
# ─────────────────────────────────────────────────────────────────────────────
def create_summary_sheet(wb):
    ws = wb.create_sheet("📊 Thống Kê", 2)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15

    # Title
    ws.merge_cells("B1:E1")
    c = ws["B1"]
    c.value = "📊 THỐNG KÊ KỊCH BẢN KIỂM THỬ ĐĂNG NHẬP"
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = make_fill(COLORS["title_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # ── Thống kê theo nhóm ──
    ws.merge_cells("B3:E3")
    c = ws["B3"]
    c.value = "THỐNG KÊ THEO NHÓM"
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = make_fill(COLORS["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    group_headers = ["NHÓM TEST", "TỔNG TC", "PASS", "FAIL"]
    for ci, h in enumerate(group_headers, start=2):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill = make_fill("2F5496")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = make_border()
    ws.row_dimensions[4].height = 22

    from collections import Counter
    group_counts = Counter(tc["group"] for tc in TEST_CASES)
    groups_ordered = [
        "Đăng nhập thành công",
        "Validation đầu vào",
        "Sai thông tin đăng nhập",
        "Bảo mật",
        "Response format",
        "Sử dụng JWT token",
        "HTTP Method & Header",
    ]
    for ri, g in enumerate(groups_ordered, start=5):
        ws.row_dimensions[ri].height = 22
        count = group_counts.get(g, 0)
        color = GROUP_COLORS.get(g, "FFFFFF")
        for ci, val in enumerate([g, count, "", ""], start=2):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = make_fill(color if ci in [2, 3] else "FFFFFF")
            c.font = Font(name="Calibri", size=10, bold=(ci == 2))
            c.alignment = Alignment(
                horizontal="center" if ci > 2 else "left",
                vertical="center"
            )
            c.border = make_border()

    # Total
    total_row = 5 + len(groups_ordered)
    ws.row_dimensions[total_row].height = 25
    for ci, val in enumerate(["TỔNG CỘNG", len(TEST_CASES), "", ""], start=2):
        c = ws.cell(row=total_row, column=ci, value=val)
        c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill = make_fill(COLORS["header_bg"])
        c.alignment = Alignment(horizontal="center" if ci > 2 else "left", vertical="center")
        c.border = make_border("medium")

    # ── Thống kê theo độ ưu tiên ──
    r_start = total_row + 2
    ws.merge_cells(f"B{r_start}:E{r_start}")
    c = ws[f"B{r_start}"]
    c.value = "THỐNG KÊ THEO ĐỘ ƯU TIÊN"
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = make_fill(COLORS["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r_start].height = 22

    priority_counts = Counter(tc["priority"] for tc in TEST_CASES)
    for i, (p, count) in enumerate([
        ("Cao", priority_counts.get("Cao", 0)),
        ("Trung bình", priority_counts.get("Trung bình", 0)),
        ("Thấp", priority_counts.get("Thấp", 0)),
    ]):
        row = r_start + 1 + i
        ws.row_dimensions[row].height = 22
        for ci, val in enumerate([p, count], start=2):
            c = ws.cell(row=row, column=ci, value=val)
            p_color = PRIORITY_COLORS.get(p, "000000")
            c.font = Font(name="Calibri", bold=(ci == 3), size=10,
                         color=p_color if ci == 2 else "000000")
            c.fill = make_fill("FFDFD5" if p == "Cao" else ("FFEFD5" if p == "Trung bình" else "DFFFDF"))
            c.alignment = Alignment(horizontal="left" if ci == 2 else "center", vertical="center")
            c.border = make_border()

    # ── Ghi chú kết quả ──
    r2 = r_start + 5
    ws.merge_cells(f"B{r2}:E{r2}")
    c = ws[f"B{r2}"]
    c.value = "QUY ƯỚC KẾT QUẢ KIỂM THỬ"
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = make_fill(COLORS["header_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r2].height = 22

    legend = [
        ("PASS",  "Kết quả thực tế khớp với kết quả mong đợi", COLORS["pass"]),
        ("FAIL",  "Kết quả thực tế KHÔNG khớp với kết quả mong đợi", COLORS["fail"]),
        ("SKIP",  "Test case bị bỏ qua (chưa test hoặc chặn bởi bug khác)", COLORS["skip"]),
        ("N/A",   "Không áp dụng trong phiên bản hiện tại", "EEEEEE"),
    ]
    for i, (status, desc, color) in enumerate(legend):
        row = r2 + 1 + i
        ws.row_dimensions[row].height = 20
        c1 = ws.cell(row=row, column=2, value=status)
        c1.font = Font(name="Calibri", bold=True, size=10)
        c1.fill = make_fill(color)
        c1.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = make_border()

        ws.merge_cells(f"C{row}:E{row}")
        c2 = ws.cell(row=row, column=3, value=desc)
        c2.font = Font(name="Calibri", size=10)
        c2.fill = make_fill("FFFFFF")
        c2.alignment = Alignment(horizontal="left", vertical="center")
        c2.border = make_border()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    output_path = "Test_Cases_Dang_Nhap.xlsx"

    wb = openpyxl.Workbook()
    # Xóa sheet mặc định
    default_sheet = wb.active
    wb.remove(default_sheet)

    create_cover_sheet(wb)
    create_testcase_sheet(wb)
    create_summary_sheet(wb)

    wb.save(output_path)
    print(f"\n✅ Đã tạo file Excel thành công: {output_path}")
    print(f"   📋 Tổng số test cases: {len(TEST_CASES)}")
    print(f"   📂 Số nhóm: 7")
    print(f"   📄 Số sheets: 3 (Trang Bìa, Test Cases, Thống Kê)")


if __name__ == "__main__":
    main()

